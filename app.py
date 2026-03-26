# app.py

import streamlit as st
from main import evaluate_response, parse_conversation
import json
import io

st.set_page_config(page_title="eval.ai", layout="wide", page_icon="🔬")

params = st.query_params
if "lang" in params:
    st.session_state.lang = params["lang"]

def format_text(text):
    if isinstance(text, list):
        items = "".join("<li>" + str(s) + "</li>" for s in text if s)
        return "<ul style='margin:0.4rem 0 0 0;padding-left:1.3rem;'>" + items + "</ul>"
    if not text or not isinstance(text, str):
        return str(text) if text else ""
    check = text
    for i in range(1, 20):
        check = check.replace(str(i) + ") ", "|SPLIT|").replace(str(i) + ". ", "|SPLIT|")
    parts = [s.strip() for s in check.split("|SPLIT|") if s.strip()]
    if len(parts) > 1:
        items = "".join("<li>" + s + "</li>" for s in parts)
        return "<ul style='margin:0.4rem 0 0 0;padding-left:1.3rem;'>" + items + "</ul>"
    sentences = text.strip().split(". ")
    sentences = [s.strip() for s in sentences if len(s.strip()) > 20]
    if len(sentences) > 1:
        items = "".join("<li>" + s + ("." if not s.endswith(".") else "") + "</li>" for s in sentences)
        return "<ul style='margin:0.4rem 0 0 0;padding-left:1.3rem;'>" + items + "</ul>"
    return text

def export_to_excel(data, lang):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    blue_fill   = PatternFill("solid", fgColor="003189")
    red_fill    = PatternFill("solid", fgColor="e53935")
    orange_fill = PatternFill("solid", fgColor="f57c00")
    green_fill  = PatternFill("solid", fgColor="2e7d32")
    gray_fill   = PatternFill("solid", fgColor="9e9e9e")
    white_font  = Font(color="FFFFFF", bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = ["Criterion", "Score", "Applicable", "Problematic Exchange", "Observed", "Justification", "Improvement Advice"] if lang == "en" else ["Critère", "Score", "Applicable", "Échange problématique", "Observé", "Justification", "Conseil"]
    col_widths = [28, 10, 12, 22, 40, 40, 40]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = white_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    for row_idx, (criterion, content_data) in enumerate(data["evaluation"].items(), 2):
        score = content_data.get("score")
        applicable = content_data.get("applicable", True)
        observed = content_data.get("observed_elements", "") or ""
        justif = content_data.get("justification", "") or ""
        advice = content_data.get("improvement_advice", "") or ""
        prob_exchange = content_data.get("problematic_exchange") or ""
        prob_detail = content_data.get("problematic_detail") or ""
        prob_combined = (prob_exchange + " — " + prob_detail) if prob_exchange and prob_detail else (prob_exchange or prob_detail or "")

        criterion_name = criterion.replace("_", " ").title()
        score_display = str(score) + " / 5" if score is not None else "N/A"
        applicable_display = ("Yes" if applicable else "No") if lang == "en" else ("Oui" if applicable else "Non")

        def to_str(v):
            if isinstance(v, list): return " | ".join(str(x) for x in v)
            return str(v) if v is not None else ""

        for col_idx, value in enumerate([criterion_name, score_display, applicable_display, prob_combined, to_str(observed), to_str(justif), to_str(advice)], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            cell.font = normal_font

        score_cell = ws.cell(row=row_idx, column=2)
        score_cell.font = Font(color="FFFFFF", bold=True, size=10)
        score_cell.alignment = Alignment(horizontal='center', vertical='center')
        if not applicable or score is None:
            score_cell.fill = gray_fill
        elif score <= 2:
            score_cell.fill = red_fill
        elif score <= 4:
            score_cell.fill = orange_fill
        else:
            score_cell.fill = green_fill

        ws.row_dimensions[row_idx].height = 80

    ws2 = wb.create_sheet(title="Suggestions" if lang == "en" else "Suggestions globales")
    ws2.column_dimensions["A"].width = 80
    suggestions = data.get("global_improvement_suggestions", [])
    title_cell = ws2.cell(row=1, column=1, value="Global Improvement Suggestions" if lang == "en" else "Suggestions d'amélioration globales")
    title_cell.font = white_font
    title_cell.fill = blue_fill
    title_cell.alignment = Alignment(wrap_text=True)
    for i, s in enumerate(suggestions, 2):
        cell = ws2.cell(row=i, column=1, value="• " + s)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = normal_font
        cell.border = thin_border
        ws2.row_dimensions[i].height = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if "lang" not in st.session_state:
    st.session_state.lang = "en"
lang = st.session_state.lang

FLAG_FR  = "https://raw.githubusercontent.com/margauxlebecque3ds-pixel/agent-evaluator/master/FR.png"
FLAG_ENG = "https://raw.githubusercontent.com/margauxlebecque3ds-pixel/agent-evaluator/master/ENG.png"

T = {
    "en": {
        "badge": "UX-driven AI evaluation",
        "title_line1": "Evaluate LEO's",
        "title_line2": "responses",
        "subtitle": "The virtual companion by Dassault Systèmes",
        "link": '<a href="https://raw.githubusercontent.com/margauxlebecque3ds-pixel/agent-evaluator/master/10heuristics.pdf" target="_blank" style="color:#4d8bff;text-decoration:none;">See the 10 heuristics list in details →</a>',
        "label_single_prompt": "USER QUESTION",
        "label_single_response": "LEO'S RESPONSE",
        "label_conversation": "PASTE CONVERSATION",
        "placeholder_single_prompt": "Paste the question asked to LEO…",
        "placeholder_single_response": "Paste LEO's response…",
        "placeholder_conversation": "Paste the full conversation here (copy/paste from SIMULIA)…",
        "tab_single": "Single exchange",
        "tab_multi": "Full conversation",
        "exchanges_detected": "exchanges detected",
        "exchange_detected": "exchange detected",
        "button": "⚡ Run Evaluation",
        "spinner": "Evaluating…",
        "results_title": "Criteria Details",
        "multi_badge": "🔁 Multi-exchange evaluation",
        "single_badge": "💬 Single exchange evaluation",
        "suggestions_title": "🌐 Global Improvement Suggestions",
        "observed": "Observed",
        "justification": "Justification",
        "problematic": "⚠️ Problematic exchange",
        "advice": "Advice",
        "na": "N/A",
        "warning": "⚠️ Please fill in the fields before running the evaluation.",
        "error": "Error",
        "export": "📥 Export to Excel",
    },
    "fr": {
        "badge": "Évaluation UX par IA",
        "title_line1": "Évaluez les réponses",
        "title_line2": "de LEO",
        "subtitle": "Le virtual companion de Dassault Systèmes",
        "link": '<a href="https://raw.githubusercontent.com/margauxlebecque3ds-pixel/agent-evaluator/master/10heuritiques.pdf" target="_blank" style="color:#4d8bff;text-decoration:none;">Voir la liste des 10 heuristiques en détail →</a>',
        "label_single_prompt": "QUESTION UTILISATEUR",
        "label_single_response": "RÉPONSE DE LEO",
        "label_conversation": "COLLER LA CONVERSATION",
        "placeholder_single_prompt": "Collez la question posée à LEO…",
        "placeholder_single_response": "Collez la réponse de LEO…",
        "placeholder_conversation": "Collez la conversation complète ici (copier/coller depuis SIMULIA)…",
        "tab_single": "Échange unique",
        "tab_multi": "Conversation complète",
        "exchanges_detected": "échanges détectés",
        "exchange_detected": "échange détecté",
        "button": "⚡ Lancer l'évaluation",
        "spinner": "Évaluation en cours…",
        "results_title": "Détail des critères",
        "multi_badge": "🔁 Évaluation multi-échanges",
        "single_badge": "💬 Évaluation échange unique",
        "suggestions_title": "🌐 Suggestions d'amélioration globales",
        "observed": "Observé",
        "justification": "Justification",
        "problematic": "⚠️ Échange problématique",
        "advice": "Conseil",
        "na": "N/A",
        "warning": "⚠️ Remplis les champs avant de lancer l'évaluation.",
        "error": "Erreur",
        "export": "📥 Exporter en Excel",
    }
}

fr_active = "background:#0a1a3a;border:1.5px solid #0050c8;color:white;" if lang=="fr" else "background:#111;border:1.5px solid #2a2a2a;color:#888;"
en_active = "background:#0a1a3a;border:1.5px solid #0050c8;color:white;" if lang=="en" else "background:#111;border:1.5px solid #2a2a2a;color:#888;"

st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Inter:wght@300;400;500;600;700&display=swap');
  .stApp {{ background-color: #0d0d0d; color: #e0e0e0; }}
  .hero {{ text-align:center; padding:2rem 1rem 2.5rem 1rem; max-width:700px; margin:0 auto; }}
  .hero-badge {{ display:inline-flex; align-items:center; gap:0.4rem; background:#111; border:1px solid #2a2a2a; border-radius:20px; padding:0.3rem 1rem; font-family:'Space Mono',monospace; font-size:0.75rem; color:#ccc; margin-bottom:1.5rem; }}
  .dot {{ width:7px; height:7px; background:#0050c8; border-radius:50%; display:inline-block; }}
  .hero-title {{ font-family:'Inter',sans-serif; font-size:3rem; font-weight:700; color:white; line-height:1.15; margin-bottom:0.3rem; }}
  .accent {{ color:#4d8bff; }}
  .hero-subtitle {{ font-family:'Inter',sans-serif; font-size:1rem; color:#aaa; margin-bottom:1rem; }}
  .hero-link {{ font-family:'Inter',sans-serif; font-size:0.85rem; color:#4d8bff; }}
  .form-label {{ font-family:'Space Mono',monospace; font-size:0.72rem; letter-spacing:0.12em; color:#aaa; text-transform:uppercase; margin-bottom:0.5rem; }}
  .exchange-badge {{ display:inline-flex; align-items:center; gap:0.4rem; padding:0.3rem 0.9rem; border-radius:20px; font-family:'Space Mono',monospace; font-size:0.75rem; font-weight:700; margin-bottom:1rem; }}
  .badge-multi {{ background:#0a1a3a; border:1px solid #0050c8; color:#4d8bff; }}
  .badge-single {{ background:#111; border:1px solid #2a2a2a; color:#888; }}
  .exchange-count {{ background:#0050c8; color:white; border-radius:10px; padding:0.1rem 0.5rem; font-size:0.7rem; margin-left:0.3rem; }}
  textarea {{ background:#111 !important; border:1px solid #2a2a2a !important; border-radius:10px !important; color:#e0e0e0 !important; font-family:'Space Mono',monospace !important; font-size:0.85rem !important; }}
  textarea::placeholder {{ color:#555 !important; }}
  textarea:focus {{ border-color:#0050c8 !important; box-shadow:0 0 0 2px rgba(0,80,200,0.15) !important; outline:none !important; }}
  /* Override Streamlit red focus on all inputs */
  [data-baseweb="input"]:focus-within,
  [data-baseweb="textarea"]:focus-within {{
    border-color:#0050c8 !important;
    box-shadow:0 0 0 2px rgba(0,80,200,0.15) !important;
  }}
  /* Mode switcher active tab underline */
  button[data-testid="baseButton-secondary"]:focus {{
    border-color:#0050c8 !important;
    box-shadow:none !important;
    outline:none !important;
  }}
  /* Red line under tabs -> blue */
  [data-testid="stMarkdownContainer"] hr {{ display:none; }}
  div[role="tab"][aria-selected="true"] {{ border-bottom:2px solid #0050c8 !important; }}
  .lang-pills {{ display:flex; gap:8px; justify-content:flex-end; padding-top:0.9rem; }}
  .lang-pill {{ display:inline-flex; align-items:center; gap:6px; padding:4px 10px 4px 6px; border-radius:20px; font-family:'Space Mono',monospace; font-size:0.72rem; font-weight:700; letter-spacing:0.05em; text-decoration:none; cursor:pointer; transition:all 0.15s; }}
  .lang-pill img {{ width:20px; height:14px; object-fit:cover; border-radius:3px; display:block; }}
  .lang-pill:hover {{ border-color:#0050c8 !important; color:white !important; }}
  .lang-pill-fr {{ {fr_active} }}
  .lang-pill-en {{ {en_active} }}
  .stButton {{ display:flex; justify-content:center; margin-top:1.5rem; }}
  .stButton > button {{ background:linear-gradient(90deg,#003189,#0050c8) !important; color:white !important; border:none !important; padding:0.75rem 2.5rem !important; border-radius:10px !important; font-family:'Space Mono',monospace !important; font-size:0.9rem !important; font-weight:700 !important; }}
  .results-title {{ font-family:'Inter',sans-serif; font-size:1.3rem; font-weight:700; color:white; margin:2rem 0 1.5rem 0; border-bottom:1px solid #1e1e1e; padding-bottom:0.8rem; }}
  .criterion-card {{ background:#111; border:1px solid #1e1e1e; border-left:4px solid #0050c8; border-radius:10px; padding:1.2rem 1.5rem; margin-bottom:1rem; }}
  .criterion-card.red {{ border-left-color:#e53935; }}
  .criterion-card.orange {{ border-left-color:#f57c00; }}
  .criterion-card.green {{ border-left-color:#2e7d32; }}
  .criterion-card.gray {{ border-left-color:#444; }}
  .crit-header {{ display:flex; flex-direction:row-reverse; justify-content:flex-end; align-items:center; gap:1rem; margin-bottom:0.8rem; }}
  .crit-name {{ font-family:'Inter',sans-serif; font-weight:600; font-size:1rem; color:#e0e0e0; }}
  .score-pill {{ font-family:'Space Mono',monospace; font-size:1.1rem; font-weight:700; padding:0.3rem 1rem; border-radius:20px; }}
  .pill-red {{ background:#2a1010; color:#e53935; }}
  .pill-orange {{ background:#2a1a00; color:#f57c00; }}
  .pill-green {{ background:#0a2a0a; color:#4caf50; }}
  .pill-gray {{ background:#1a1a1a; color:#777; }}
  .crit-detail {{ font-family:'Inter',sans-serif; font-size:0.92rem; color:#c0c0c0; margin-top:0.6rem; line-height:1.7; }}
  .crit-detail strong {{ color:#e0e0e0; }}
  .crit-detail ul {{ margin:0.4rem 0 0 0; padding-left:1.3rem; }}
  .crit-detail li {{ margin-bottom:0.3rem; }}
  .problematic-box {{ background:#1a0a0a; border:1px solid #e53935; border-radius:6px; padding:0.5rem 0.8rem; margin-top:0.5rem; font-family:'Space Mono',monospace; font-size:0.78rem; color:#e57373; }}
  .suggestions-box {{ background:#0a0f1a; border:1px solid #1a2a4a; border-radius:10px; padding:1.5rem; margin-top:2rem; }}
  .suggestions-box h3 {{ font-family:'Inter',sans-serif; color:#4d8bff; font-size:1rem; margin-bottom:1rem; }}
  .suggestions-box li {{ font-family:'Inter',sans-serif; font-size:0.92rem; color:#c0c0c0; margin-bottom:0.5rem; line-height:1.7; }}
  div[data-testid="stTabs"] button {{ font-family:'Space Mono',monospace !important; font-size:0.78rem !important; color:#aaa !important; }}
  div[data-testid="stTabs"] button[aria-selected="true"] {{ color:white !important; border-bottom:2px solid #0050c8 !important; }}
  #MainMenu, footer, header {{ visibility:hidden; }}
  .block-container {{ padding-top:0 !important; }}
  label {{ display:none !important; }}
  hr {{ display:none; }}

  /* Override Streamlit red accent with blue */
  .stTabs [data-baseweb="tab-highlight"] {{ background-color: #0050c8 !important; }}
  .stTabs [data-baseweb="tab"] {{ color: #aaa !important; font-family:'Space Mono',monospace !important; font-size:0.8rem !important; }}
  .stTabs [aria-selected="true"] {{ color: white !important; }}
  textarea:focus {{ border-color: #0050c8 !important; box-shadow: 0 0 0 2px rgba(0,80,200,0.15) !important; outline: none !important; }}
  input:focus {{ border-color: #0050c8 !important; box-shadow: 0 0 0 2px rgba(0,80,200,0.15) !important; }}
  [data-baseweb="input"]:focus-within {{ border-color: #0050c8 !important; }}
  :root {{ --primary: #0050c8 !important; }}
  .stButton > button:focus {{ border-color: #0050c8 !important; box-shadow: 0 0 0 2px rgba(0,80,200,0.15) !important; }}
</style>
""", unsafe_allow_html=True)

# Navbar
st.markdown(f"""
<div style="display:flex;align-items:center;justify-content:space-between;padding:0.8rem 0;border-bottom:1px solid #1e1e1e;margin-bottom:2rem;">
  <img src="https://raw.githubusercontent.com/margauxlebecque3ds-pixel/agent-evaluator/master/DassaultSyst%C3%A8mesLogo.png" style="height:36px;width:auto;" />
  <div class="lang-pills">
    <a href="?lang=fr" target="_self" class="lang-pill lang-pill-fr">
      <img src="{FLAG_FR}" alt="FR"/> FR
    </a>
    <a href="?lang=en" target="_self" class="lang-pill lang-pill-en">
      <img src="{FLAG_ENG}" alt="EN"/> EN
    </a>
  </div>
</div>
""", unsafe_allow_html=True)

t = T[lang]

st.markdown(f"""
<div class="hero">
  <div class="hero-badge"><span class="dot"></span> {t["badge"]}</div>
  <div class="hero-title">{t["title_line1"]}<br><span class="accent">{t["title_line2"]}</span></div>
  <div class="hero-subtitle">{t["subtitle"]}</div>
  <span class="hero-link">{t["link"]}</span>
</div>
""", unsafe_allow_html=True)

# Tabs
tab1, tab2 = st.tabs([t["tab_single"], t["tab_multi"]])

with tab1:
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f'<div class="form-label">{t["label_single_prompt"]}</div>', unsafe_allow_html=True)
        single_prompt = st.text_area("q", height=180, placeholder=t["placeholder_single_prompt"], label_visibility="collapsed")
    with col2:
        st.markdown(f'<div class="form-label">{t["label_single_response"]}</div>', unsafe_allow_html=True)
        single_response = st.text_area("r", height=180, placeholder=t["placeholder_single_response"], label_visibility="collapsed")

    if st.button(t["button"], key="btn_single"):
        if single_prompt and single_response:
            with st.spinner(t["spinner"]):
                evaluation = evaluate_response(single_prompt, single_response, language=lang)
            try:
                from json_repair import repair_json
                data = json.loads(repair_json(evaluation))
                st.session_state.last_results = data
                st.session_state.last_lang = lang
                st.session_state.last_mode = "single"
            except Exception as e:
                st.error(f"{t['error']} : {e}")
                st.code(evaluation)
                st.stop()
        else:
            st.warning(t["warning"])

with tab2:
    st.markdown(f'<div class="form-label">{t["label_conversation"]}</div>', unsafe_allow_html=True)

    conv_input = st.text_area("conv", height=280, placeholder=t["placeholder_conversation"], label_visibility="collapsed")

    # Live detection badge
    if conv_input:
        pairs = parse_conversation(conv_input)
        if pairs and len(pairs) > 1:
            count = len(pairs)
            st.markdown(f'<div class="exchange-badge badge-multi">{t["multi_badge"]} <span class="exchange-count">{count} {t["exchanges_detected"]}</span></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="exchange-badge badge-single">{t["single_badge"]}</div>', unsafe_allow_html=True)

    if st.button(t["button"], key="btn_multi"):
        if conv_input:
            with st.spinner(t["spinner"]):
                evaluation = evaluate_response("", "", language=lang, mode="multi", conversation_raw=conv_input)
            try:
                from json_repair import repair_json
                data = json.loads(repair_json(evaluation))
                st.session_state.last_results = data
                st.session_state.last_lang = lang
                st.session_state.last_mode = "multi"
            except Exception as e:
                st.error(f"{t['error']} : {e}")
                st.code(evaluation)
                st.stop()
        else:
            st.warning(t["warning"])

# Display results
if "last_results" in st.session_state and st.session_state.last_results:
    data = st.session_state.last_results
    res_lang = st.session_state.get("last_lang", lang)
    res_t = T[res_lang]
    is_multi = data.get("is_multi_exchange", False)
    num_exchanges = data.get("num_exchanges", 1)

    if is_multi:
        st.markdown(f'<div class="exchange-badge badge-multi">{res_t["multi_badge"]} <span class="exchange-count">{num_exchanges} {res_t["exchanges_detected"]}</span></div>', unsafe_allow_html=True)

    n_ex = st.session_state.get("last_n_exchanges", 1)
    mode_badge = f" &nbsp;<span style='font-size:0.75rem;background:#0a1a3a;border:1px solid #0050c8;color:#4d8bff;padding:2px 10px;border-radius:20px;font-family:Space Mono,monospace;'>{n_ex} exchange{'s' if n_ex > 1 else ''}</span>" if n_ex > 1 else ""
    st.markdown(f'<div class="results-title">{res_t["results_title"]}{mode_badge}</div>', unsafe_allow_html=True)

    for criterion, content_item in data["evaluation"].items():
        criterion_name = criterion.replace("_", " ").title()
        score = content_item.get("score")
        applicable = content_item.get("applicable", True)
        if not applicable or score is None:
            color = "gray"; display_score = res_t["na"]; pill = "pill-gray"
        elif score <= 2:
            color = "red"; display_score = f"{score} / 5"; pill = "pill-red"
        elif score <= 4:
            color = "orange"; display_score = f"{score} / 5"; pill = "pill-orange"
        else:
            color = "green"; display_score = f"{score} / 5"; pill = "pill-green"

        observed = content_item.get("observed_elements", "") or ""
        justif   = content_item.get("justification", "") or ""
        advice   = content_item.get("improvement_advice", "") or ""
        prob_exchange = content_item.get("problematic_exchange") or ""
        prob_detail   = content_item.get("problematic_detail") or ""

        html = f'<div class="criterion-card {color}"><div class="crit-header"><div class="crit-name">{criterion_name}</div><div class="score-pill {pill}">{display_score}</div></div>'
        if observed:
            html += f'<div class="crit-detail"><strong>🔍 {res_t["observed"]}:</strong> {format_text(observed)}</div>'
        if justif:
            html += f'<div class="crit-detail"><strong>📋 {res_t["justification"]}:</strong> {format_text(justif)}</div>'
        if prob_exchange:
            html += f'<div class="problematic-box">⚠️ {res_t["problematic"]}: {prob_exchange}'
            if prob_detail:
                html += f' — {prob_detail}'
            html += '</div>'
        if advice and applicable and score is not None and score < 5:
            html += f'<div class="crit-detail"><strong>💡 {res_t["advice"]}:</strong> {format_text(advice)}</div>'
        html += "</div>"
        st.markdown(html, unsafe_allow_html=True)

    suggestions = data.get("global_improvement_suggestions", [])
    if suggestions:
        s_html = f'<div class="suggestions-box"><h3>{res_t["suggestions_title"]}</h3><ul>'
        for s in suggestions:
            s_html += f"<li>{s}</li>"
        s_html += "</ul></div>"
        st.markdown(s_html, unsafe_allow_html=True)

    excel_file = export_to_excel(data, res_lang)
    st.download_button(
        label=res_t["export"],
        data=excel_file,
        file_name="evaluation_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )